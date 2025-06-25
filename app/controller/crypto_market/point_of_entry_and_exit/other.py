from openpyxl import Workbook, load_workbook
import os
import numpy as np


def create_sequences(data, seq_length=50):
    xs, ys = [], []
    for i in range(len(data) - seq_length):
        x = data[i:i + seq_length]
        y = data[i + seq_length]
        xs.append(x)
        ys.append(y)
    return np.array(xs), np.array(ys)


def log_training_to_excel(filename, row_data):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Training Results"
        headers = ["symbol", "model", "optimizer", "data_size", "window_size", "epochs",
                   "batch_size", "avg loss", "mae", "rmse", "training_time"]
        ws.append(headers)
    else:
        wb = load_workbook(filename)
        ws = wb["Training Results"]

    # Перевіряємо, чи існує такий рядок
    symbol, model, optimizer, data_size, window_size, epochs, batch_size = row_data[:7]

    for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаємо перший рядок (заголовки)
        if (row[0] == symbol and row[1] == model and row[2] == optimizer and
                row[3] == data_size and row[4] == window_size and row[5] == epochs and row[6] == batch_size):
            return  # Якщо знайдено такий самий запис, не додаємо новий

    # Якщо запис не знайдений, додаємо новий
    ws.append(row_data)
    wb.save(filename)
